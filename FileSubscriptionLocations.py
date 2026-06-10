"""
FileSubscriptionLocations.py

Export and update file-subscription delivery locations.

Two-phase workflow
──────────────────
  export  Writes an Excel workbook with two sheets:

          Devices       — every FILE-type device; columns cover all FileLocation
                          and FileProperties attributes.
                          Fill in NewFilePath to change the device's base
                          delivery folder.

          UserAddresses — every user address whose device has
                          append_user_path=True (i.e. user-controlled paths).
                          Fill in NewPhysicalAddress to relocate a user's
                          delivery folder.

  apply   Reads the workbook produced by export, applies every non-empty
          NewFilePath / NewPhysicalAddress value to the live environment.
          Dry-run by default; pass --apply to commit changes.

Connection
──────────
  Default     : mstrio_core / .env  (CLI / PyCharm)
  Workstation : search "WORKSTATION" below to swap.

Usage
─────
  python FileSubscriptionLocations.py dev export
  python FileSubscriptionLocations.py dev export --output-dir C:/tmp
  python FileSubscriptionLocations.py dev apply --input C:/tmp/file_sub_locations_dev_20260610_120000.xlsx
  python FileSubscriptionLocations.py dev apply --input ... --apply
  python FileSubscriptionLocations.py dev export --concurrency 20
"""

import argparse
import concurrent.futures
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from loguru import logger

from mstrio.distribution_services import Device, DeviceType, FileDeviceProperties, list_devices
from mstrio.users_and_groups import User, list_users

from mstrio_core import MstrConfig, get_mstrio_connection
from mstrio_core.config import MstrEnvironment

# ── Workstation import (uncomment when running from Workstation) ──────────────
# from mstrio.connection import get_connection

# ══════════════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════

OUTPUT_FILENAME     = 'file_sub_locations_{env}_{ts}.xlsx'
DEFAULT_CONCURRENCY = 10

DEVICES_SHEET   = 'Devices'
ADDRESSES_SHEET = 'UserAddresses'

# Columns in the Devices sheet.
# file_path / append_user_path come from FileLocation;
# read_only … unix_access_rights come from FileProperties.
DEVICE_COLS = [
    'device_id',
    'device_name',
    'file_path',
    'append_user_path',
    'use_backup_location',
    'backup_file_path',
    'read_only',
    'archive',
    'index',
    'file_encoding',
    'unix_access_rights',
    'NewFilePath',          # user fills this in
]

ADDRESS_COLS = [
    'user_id',
    'user_name',
    'login_id',
    'address_id',
    'address_name',
    'device_id',
    'device_name',
    'physical_address',
    'NewPhysicalAddress',   # user fills this in
]

_NEW_FILL = PatternFill('solid', fgColor='FFF2CC')  # light yellow highlight
_HDR_FONT = Font(bold=True)


# ══════════════════════════════════════════════════════════════════════════════
#  DEVICE HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _device_row(device: Device) -> dict:
    """Flatten a FILE-type Device into a dict matching DEVICE_COLS."""
    row: dict = {col: '' for col in DEVICE_COLS}
    row['device_id']   = device.id
    row['device_name'] = device.name or ''

    dp = device.device_properties
    if not isinstance(dp, FileDeviceProperties):
        return row  # non-FILE device — leave file fields blank

    fl = dp.file_location
    if fl:
        row['file_path']           = fl.file_path           or ''
        row['append_user_path']    = fl.append_user_path
        row['use_backup_location'] = fl.use_backup_location
        row['backup_file_path']    = fl.backup_file_path    or ''

    fp = dp.file_properties
    if fp:
        row['read_only']          = fp.read_only
        row['archive']            = fp.archive
        row['index']              = fp.index
        fe = fp.file_encoding
        row['file_encoding']      = fe.value if hasattr(fe, 'value') else (str(fe) if fe else '')
        row['unix_access_rights'] = fp.unix_access_rights  or ''

    return row


# ══════════════════════════════════════════════════════════════════════════════
#  USER-ADDRESS HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _user_address_rows(conn, user_stub, append_device_ids: set[str]) -> list[dict]:
    """
    Fetch the full User and return one row per ContactAddress whose device
    is in append_device_ids.  Returns [] on failure (already logged).

    list_users() returns lightweight objects without addresses populated;
    User(conn, id=) forces a full fetch including the addresses list.
    """
    try:
        user = User(conn, id=user_stub.id)
        addresses = user.addresses or []
    except Exception as exc:
        logger.warning('Cannot fetch addresses for user {id}: {exc}', id=user_stub.id, exc=exc)
        return []

    login_id = getattr(user, 'username', '') or ''

    rows = []
    for addr in addresses:
        dev = getattr(addr, 'device', None)
        if dev is None:
            continue
        if dev.id.upper() not in append_device_ids:
            continue

        rows.append({
            'user_id':          user.id,
            'user_name':        user.name          or '',
            'login_id':         login_id,
            'address_id':       addr.id            or '',
            'address_name':     addr.name          or '',
            'device_id':        dev.id             or '',
            'device_name':      dev.name           or '',
            'physical_address': addr.physical_address or '',
            'NewPhysicalAddress': '',
        })

    return rows


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _write_sheet(ws, columns: list[str], rows: list[dict], new_col: str) -> None:
    """Write a bold header row + data rows; highlight the user-input column."""
    new_col_idx = columns.index(new_col) + 1  # 1-based

    # Header
    for ci, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = _HDR_FONT
        if ci == new_col_idx:
            cell.fill = _NEW_FILL

    # Data
    for ri, row in enumerate(rows, 2):
        for ci, col in enumerate(columns, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(col, ''))
            if ci == new_col_idx:
                cell.fill = _NEW_FILL

    # Auto-width (approximate)
    for ci, col in enumerate(columns, 1):
        max_len = max(
            len(col),
            max((len(str(r.get(col) or '')) for r in rows), default=0),
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 60)

    ws.freeze_panes = 'A2'


def _read_sheet(wb: openpyxl.Workbook, name: str) -> list[dict]:
    """Return all data rows from a sheet as list-of-dicts keyed by header."""
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        result.append({
            headers[i]: (row[i] if i < len(row) and row[i] is not None else '')
            for i in range(len(headers))
        })
    return result


# ══════════════════════════════════════════════════════════════════════════════
#  COMMANDS
# ══════════════════════════════════════════════════════════════════════════════

def cmd_export(conn, env: str, out_dir: Path, concurrency: int) -> int:
    # ── 1. Load all FILE devices ──────────────────────────────────────────────
    logger.info('Loading FILE devices...')
    devices     = list_devices(conn, device_type=DeviceType.FILE)
    device_rows = [_device_row(d) for d in devices]
    logger.info('{n} file device(s) loaded', n=len(device_rows))

    # Build the set of device IDs where append_user_path=True
    append_device_ids: set[str] = {
        r['device_id'].upper()
        for r in device_rows
        if r.get('append_user_path') is True
    }
    logger.info(
        '{n} device(s) have append_user_path=True',
        n=len(append_device_ids),
    )

    # ── 2. Load users → collect matching addresses in parallel ────────────────
    if append_device_ids:
        logger.info('Loading users...')
        user_stubs = list_users(conn)
        logger.info('{n} user(s) — fetching addresses...', n=len(user_stubs))

        address_rows: list[dict] = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=concurrency) as pool:
            futures = {
                pool.submit(_user_address_rows, conn, u, append_device_ids): u
                for u in user_stubs
            }
            done = 0
            for fut in concurrent.futures.as_completed(futures):
                done += 1
                if done % 200 == 0 or done == len(futures):
                    logger.info('  {d}/{t} users processed', d=done, t=len(futures))
                try:
                    address_rows.extend(fut.result())
                except Exception as exc:
                    user = futures[fut]
                    logger.warning('Error processing user {id}: {exc}', id=user.id, exc=exc)

        logger.info('{n} address row(s) on append_user_path devices', n=len(address_rows))
    else:
        logger.warning('No devices with append_user_path=True — UserAddresses sheet will be empty.')
        address_rows = []

    # ── 3. Write workbook ─────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop default empty sheet

    _write_sheet(wb.create_sheet(DEVICES_SHEET),   DEVICE_COLS,  device_rows,  'NewFilePath')
    _write_sheet(wb.create_sheet(ADDRESSES_SHEET), ADDRESS_COLS, address_rows, 'NewPhysicalAddress')

    ts       = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_file = out_dir / OUTPUT_FILENAME.format(env=env, ts=ts)
    out_dir.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_file))

    logger.success(
        'Workbook written → {p}  ({d} device rows, {a} address rows)',
        p=out_file, d=len(device_rows), a=len(address_rows),
    )
    print(f'\nExported: {out_file}')
    return 0


def cmd_apply(conn, input_path: Path, do_apply: bool) -> int:
    wb = openpyxl.load_workbook(str(input_path))

    dev_updated = dev_skipped = dev_errors = 0
    addr_updated = addr_skipped = addr_errors = 0

    # ── Devices ───────────────────────────────────────────────────────────────
    for row in _read_sheet(wb, DEVICES_SHEET):
        device_id    = str(row.get('device_id')   or '').strip()
        new_path     = str(row.get('NewFilePath') or '').strip()
        current_path = str(row.get('file_path')  or '').strip()

        if not device_id or not new_path:
            continue
        if new_path == current_path:
            dev_skipped += 1
            continue

        logger.info(
            'Device {id} ({name}): {old!r} → {new!r}',
            id=device_id, name=row.get('device_name', ''),
            old=current_path, new=new_path,
        )
        if do_apply:
            try:
                device = Device(conn, id=device_id)
                dp = device.device_properties
                dp.file_location.file_path = new_path
                device.alter(device_properties=dp)
                dev_updated += 1
                logger.success('  updated device {id}', id=device_id)
            except Exception as exc:
                dev_errors += 1
                logger.error('  ERROR updating device {id}: {exc}', id=device_id, exc=exc)
        else:
            logger.info('  [DRY-RUN] would update device {id}', id=device_id)
            dev_updated += 1

    # ── UserAddresses ─────────────────────────────────────────────────────────
    for row in _read_sheet(wb, ADDRESSES_SHEET):
        user_id      = str(row.get('user_id')            or '').strip()
        address_id   = str(row.get('address_id')         or '').strip()
        new_path     = str(row.get('NewPhysicalAddress') or '').strip()
        current_path = str(row.get('physical_address')   or '').strip()

        if not user_id or not address_id or not new_path:
            continue
        if new_path == current_path:
            addr_skipped += 1
            continue

        logger.info(
            'User {uid} ({uname}) addr {aid}: {old!r} → {new!r}',
            uid=user_id, uname=row.get('user_name', ''),
            aid=address_id, old=current_path, new=new_path,
        )
        if do_apply:
            try:
                user = User(conn, id=user_id)
                # Note: update_address parameter is 'address', not 'physical_address'
                user.update_address(id=address_id, address=new_path)
                addr_updated += 1
                logger.success('  updated address {aid}', aid=address_id)
            except Exception as exc:
                addr_errors += 1
                logger.error(
                    '  ERROR updating address {aid} for user {uid}: {exc}',
                    aid=address_id, uid=user_id, exc=exc,
                )
        else:
            logger.info('  [DRY-RUN] would update address {aid} for user {uid}', aid=address_id, uid=user_id)
            addr_updated += 1

    label = 'Applied' if do_apply else 'Dry-run'
    summary = (
        f'\n{label} — '
        f'Devices: {dev_updated} updated, {dev_skipped} skipped, {dev_errors} errors  |  '
        f'Addresses: {addr_updated} updated, {addr_skipped} skipped, {addr_errors} errors'
    )
    print(summary)
    logger.info(summary.strip())
    return 0 if (dev_errors + addr_errors) == 0 else 1


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            'Export and update file-subscription delivery locations. '
            'Run "export" to generate the workbook, fill in the New* columns, '
            'then run "apply --apply" to commit changes.'
        )
    )
    parser.add_argument(
        'env',
        choices=[e.value for e in MstrEnvironment],
        help='MicroStrategy environment (dev / qa / prod).',
    )

    sub = parser.add_subparsers(dest='command', required=True)

    p_exp = sub.add_parser('export', help='Export devices and user addresses to Excel.')
    p_exp.add_argument('--output-dir', type=Path, default=None, metavar='DIR',
                       help='Output directory (default: MSTR_OUTPUT_DIR).')
    p_exp.add_argument('--concurrency', type=int, default=DEFAULT_CONCURRENCY, metavar='N',
                       help=f'Parallel user-address fetches (default: {DEFAULT_CONCURRENCY}).')

    p_app = sub.add_parser('apply', help='Apply NewFilePath / NewPhysicalAddress changes.')
    p_app.add_argument('--input', '-i', required=True, type=Path, metavar='XLSX',
                       help='Workbook produced by the export subcommand.')
    p_app.add_argument('--apply', dest='do_apply', action='store_true',
                       help='Commit changes (default: dry-run).')

    args   = parser.parse_args()
    config = MstrConfig(environment=MstrEnvironment(args.env))

    # ── Option A: mstrio_core / .env (CLI / PyCharm) ─────────────────────────
    conn = get_mstrio_connection(config=config)

    # ── Option B: Workstation (WORKSTATION) ───────────────────────────────────
    # Comment out Option A above and uncomment below.
    # conn = get_connection(workstationData)  # noqa: F821

    if args.command == 'export':
        out_dir = args.output_dir or config.output_dir
        return cmd_export(conn, args.env, out_dir, args.concurrency)

    if args.command == 'apply':
        if not args.input.exists():
            logger.error('Input file not found: {p}', p=args.input)
            return 1
        if not args.do_apply:
            logger.info('DRY-RUN mode — pass --apply to commit changes.')
        return cmd_apply(conn, args.input, args.do_apply)

    return 0


if __name__ == '__main__':
    raise SystemExit(main())
